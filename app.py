#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Flask Web Server for XML to Excel Conversion
Complete implementation ready for deployment
"""

from flask import Flask, request, send_file, send_from_directory
from flask_cors import CORS
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl import Workbook
import re
import io
import os

app = Flask(__name__)
CORS(app)

# Configuration
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

@app.route('/')
def index():
    """Serve the main HTML page"""
    return send_from_directory('.', 'index_beautiful.html')

@app.route('/api/convert', methods=['POST'])
def convert_xml_to_excel():
    """Convert XML file to Excel"""
    try:
        if 'file' not in request.files:
            return 'No file uploaded', 400
        
        file = request.files['file']
        
        if file.filename == '':
            return 'No file selected', 400
        
        if not file.filename.endswith('.xml'):
            return 'Please upload an XML file', 400
        
        # Read XML content
        xml_content = file.read().decode('utf-8')
        
        # Parse and convert
        data_rows = parse_xml_invoice(xml_content)
        
        # Create Excel in memory
        excel_file = create_excel_in_memory(data_rows)
        
        # Generate output filename
        output_filename = file.filename.replace('.xml', '.xlsx')
        
        # Return Excel file
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=output_filename
        )
        
    except Exception as e:
        return f'Error: {str(e)}', 500


def parse_customer_name(full_name):
    """Extract customer name and tax number from full name"""
    if not full_name:
        return '', None
    
    full_name = full_name.replace('&quot;', '"').replace('&amp;', '&')
    match = re.search(r'\s+(\d+)\s*$', full_name)
    
    if match:
        tax_number = match.group(1)
        name = full_name[:match.start()].strip()
        return name, tax_number
    
    return full_name.strip(), None


def parse_description(description, transaction_type):
    """Parse description to extract sub-customer, fee type, and number"""
    if not description:
        return None, None, None
    
    description = description.replace('&quot;', '"').replace('&amp;', '&')
    parts = description.split(':', 1)
    
    if len(parts) != 2:
        return None, None, None
    
    sub_customer = parts[0].strip()
    rest = parts[1].strip()
    
    fee_type = 'חיוב חודשי' if 'Recurring fee' in rest else 'חיוב חד פעמי'
    
    number_match = re.search(r'(\d+)\s*$', rest)
    number_id = number_match.group(1) if number_match else ''
    
    return sub_customer, fee_type, number_id


def parse_xml_invoice(xml_content):
    """Parse XML invoice and extract data"""
    root = ET.fromstring(xml_content)
    invoice = root.find('.//invoice')
    
    if invoice is None:
        raise ValueError("No invoice element found in XML")
    
    # Extract customer information
    customer = invoice.find('.//customer')
    customer_name_elem = customer.find('name')
    customer_full_name = customer_name_elem.text if customer_name_elem is not None else ''
    customer_name, tax_number_from_name = parse_customer_name(customer_full_name)
    
    # Get tax number
    tax_number_elem = customer.find('tax_number')
    tax_number = None
    
    if tax_number_elem is not None and tax_number_elem.text and tax_number_elem.text.strip():
        tax_number = tax_number_elem.text.strip()
    else:
        tax_number = tax_number_from_name
    
    data_rows = []
    
    # Process transactions
    transactions = invoice.find('.//transactions')
    if transactions is not None:
        for transaction in transactions.findall('transaction'):
            trans_type = transaction.find('type').text if transaction.find('type') is not None else ''
            description = transaction.find('description').text if transaction.find('description') is not None else ''
            amount_ex_tax = transaction.find('amount_ex_tax').text if transaction.find('amount_ex_tax') is not None else '0'
            
            sub_customer, fee_type, number_id = parse_description(description, trans_type)
            
            if sub_customer and fee_type and number_id:
                # Reverse sub-customer name
                sub_parts = sub_customer.split(None, 1)
                if len(sub_parts) == 2:
                    sub_customer = f"{sub_parts[1]} {sub_parts[0]}"
                
                description_text = f"{number_id}\n{fee_type}\n{sub_customer}"
                product_name = 'מספר טלפון' if trans_type == 'fees_number' else 'רשיון שלוחה'
                
                data_rows.append({
                    'customer_name': customer_name,
                    'customer_external_id': tax_number,
                    'product_name': product_name,
                    'description': description_text,
                    'quantity': 1,
                    'price': float(amount_ex_tax),
                    'billing_cycles': 1,
                    'document_only': 'כן',
                    'document_type': 'חשבון עסקה'
                })
    
    # Process calls
    calls = invoice.find('.//calls')
    if calls is not None:
        for category in calls.findall('category'):
            direction = category.get('direction')
            if direction != 'out':
                continue
            
            for group in category.findall('group'):
                group_name_elem = group.find('name')
                if group_name_elem is None:
                    continue
                
                group_name = group_name_elem.text if group_name_elem.text else ''
                talk_time_elem = group.find('talk_time')
                cost_ex_tax_elem = group.find('cost_ex_tax')
                
                if talk_time_elem is None or cost_ex_tax_elem is None:
                    continue
                
                talk_time = talk_time_elem.text if talk_time_elem.text else '0:00'
                cost = cost_ex_tax_elem.text if cost_ex_tax_elem.text else '0'
                
                product_name = f"דקות שיחה | {group_name}"
                
                data_rows.append({
                    'customer_name': customer_name,
                    'customer_external_id': tax_number,
                    'product_name': product_name,
                    'description': talk_time,
                    'quantity': 1,
                    'price': float(cost),
                    'billing_cycles': 1,
                    'document_only': 'כן',
                    'document_type': 'חשבון עסקה'
                })
    
    return data_rows


def create_excel_in_memory(data_rows):
    """Create Excel file in memory and return as BytesIO"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'חשבונית - שורות'
    
    # Headers
    headers = [
        'מזהה לקוח', 'שם מלא', 'מזהה לקוח חיצוני', 'טלפון', 'דואר אלקטרוני',
        'חודש תוקף כרטיס אשראי', 'שנת תוקף כרטיס אשראי', 'מספר כרטיס אשראי',
        'תעודת זהות בעל כרטיס', 'מספר בנק', 'מספר סניף בנק', 'מספר חשבון בנק',
        'תעודת זהות חשבון בנק', 'שם בעל החשבון', 'מזהה מוצר', 'שם מוצר',
        'תיאור', 'מזהה מוצר חיצוני', 'תדירות מוצר בחודשים', 'כמות',
        'מחיר ליחידה', 'תאריך החיוב הבא', 'מחזורי חיוב', 'מסמך בלבד',
        'סוג המסמך שיופק'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Data rows
    for row_idx, data in enumerate(data_rows, start=2):
        ws.cell(row=row_idx, column=2, value=data['customer_name'])
        ws.cell(row=row_idx, column=3, value=data['customer_external_id'])
        ws.cell(row=row_idx, column=16, value=data['product_name'])
        
        # Description with wrap text
        description_cell = ws.cell(row=row_idx, column=17, value=data['description'])
        description_cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
        
        ws.cell(row=row_idx, column=19, value=1)  # תדירות
        ws.cell(row=row_idx, column=20, value=data['quantity'])
        ws.cell(row=row_idx, column=21, value=data['price'])
        ws.cell(row=row_idx, column=23, value=data['billing_cycles'])
        ws.cell(row=row_idx, column=24, value=data['document_only'])
        ws.cell(row=row_idx, column=25, value=data['document_type'])
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


if __name__ == '__main__':
    # For development
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=True, host='0.0.0.0', port=port)
